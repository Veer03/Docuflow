import io
import os
import shutil
import zipfile
import pandas as pd
import openpyxl
import qrcode
from copy import copy
from io import BytesIO
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from docx2pdf import convert

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "https://docuflow-woad.vercel.app"
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

TEMP_DIR = "temp_files"
os.makedirs(TEMP_DIR, exist_ok=True)


def remove_file(path: str):
    try:
        if os.path.exists(path):
            os.remove(path)
            print(f"Cleaned up: {path}")
    except Exception as e:
        print(f"Cleanup failed for {path}: {str(e)}")


def write_group_to_buffer(source_ws, group_df, header_row=0):
    """Copies rows preserving original cell formatting, colors, number formats."""
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active

    # copy header row(s) with original styles
    for r in range(1, header_row + 2):
        src_row = list(source_ws.iter_rows(min_row=r, max_row=r))[0]
        for col_idx, cell in enumerate(src_row, 1):
            new_cell = new_ws.cell(row=r, column=col_idx, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format

    # copy column widths
    for col_letter, dim in source_ws.column_dimensions.items():
        new_ws.column_dimensions[col_letter].width = dim.width

    # write data rows
    data_start_row = header_row + 2
    for row_offset, (_, data_row) in enumerate(group_df.iterrows()):
        new_row_idx = data_start_row + row_offset
        for col_idx, value in enumerate(data_row, 1):
            src_cell = source_ws.cell(row=new_row_idx, column=col_idx)
            new_cell = new_ws.cell(row=new_row_idx, column=col_idx, value=value)
            if src_cell.has_style:
                new_cell.font = copy(src_cell.font)
                new_cell.fill = copy(src_cell.fill)
                new_cell.border = copy(src_cell.border)
                new_cell.alignment = copy(src_cell.alignment)
                new_cell.number_format = src_cell.number_format

    buffer = BytesIO()
    new_wb.save(buffer)
    buffer.seek(0)
    return buffer


@app.post("/api/inspect")
async def inspect_excel(
    file: UploadFile = File(...),
    sheet: str = Form(None),
    header_row: int = Form(0)
):
    contents = await file.read()
    try:
        xl = pd.ExcelFile(BytesIO(contents))
        sheets = xl.sheet_names
        target_sheet = sheet if sheet else sheets[0]
        df = xl.parse(target_sheet, header=header_row)
        columns = list(df.columns)
        return {"sheets": sheets, "columns": columns}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")


@app.post("/api/split")
async def split_excel(
    file: UploadFile = File(...),
    split_column: str = Form(...),
    sheet: str = Form(None),
    header_row: int = Form(0),
    header_color: str = Form("#a855f7")
):
    contents = await file.read()
    split_column = split_column.strip()

    try:
        xl = pd.ExcelFile(BytesIO(contents))
        target_sheet = sheet if sheet else xl.sheet_names[0]
        df = pd.read_excel(
            BytesIO(contents),
            sheet_name=target_sheet,
            header=header_row,
            engine='openpyxl'
        )
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    if split_column not in df.columns:
        raise HTTPException(
            status_code=400,
            detail=f"Column '{split_column}' not found. Available: {list(df.columns)}"
        )

    # load source workbook to copy styles from
    source_wb = openpyxl.load_workbook(BytesIO(contents))
    source_ws = source_wb[target_sheet]

    grouped = df.groupby(split_column)
    unique_keys = list(grouped.groups.keys())

    # single group — return as direct file download
    if len(unique_keys) == 1:
        key = unique_keys[0]
        group_df = grouped.get_group(key)
        final_buffer = write_group_to_buffer(source_ws, group_df, header_row)
        return StreamingResponse(
            final_buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={key}.xlsx"}
        )

    # multiple groups — zip them all
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for group_name, group_df in grouped:
            safe_name = str(group_name).replace("/", "_").replace("\\", "_")
            final_buffer = write_group_to_buffer(source_ws, group_df, header_row)
            zip_file.writestr(f"{safe_name}.xlsx", final_buffer.getvalue())

    zip_buffer.seek(0)
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=split_files.zip"}
    )


@app.post("/api/columns")
async def get_excel_columns(file: UploadFile = File(...)):
    contents = await file.read()
    try:
        df = pd.read_excel(BytesIO(contents), engine='openpyxl', nrows=0)
        return {"columns": list(df.columns)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")


@app.post("/api/word-to-pdf")
async def word_to_pdf(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...)
):
    if not file.filename.endswith(".docx"):
        raise HTTPException(status_code=400, detail="Please upload a valid .docx file.")

    input_docx_path = os.path.join(TEMP_DIR, file.filename)
    output_pdf_path = os.path.join(TEMP_DIR, file.filename.replace(".docx", ".pdf"))

    try:
        with open(input_docx_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        convert(input_docx_path, output_pdf_path)
        background_tasks.add_task(remove_file, output_pdf_path)
        return FileResponse(
            path=output_pdf_path,
            media_type="application/pdf",
            filename=os.path.basename(output_pdf_path)
        )
    except Exception as e:
        if os.path.exists(output_pdf_path):
            os.remove(output_pdf_path)
        raise HTTPException(status_code=500, detail=f"Conversion failed: {str(e)}")
    finally:
        if os.path.exists(input_docx_path):
            os.remove(input_docx_path)

@app.post("/api/generate-qr")
async def generate_qr(
    url: str = Form(...), 
    filename: str = Form("qrcode.png")
):
    try:
        # Sanitize filename to prevent directory traversal if used in headers
        safe_filename = "".join(c for c in filename if c.isalnum() or c in "._-")
        if not safe_filename.endswith(".png"):
            safe_filename += ".png"

        # Generate QR code
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=4
        )
        qr.add_data(url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
    
        # Save to memory instead of disk
        buffer = io.BytesIO()
        img.save(buffer, format="PNG")
        buffer.seek(0)

        return StreamingResponse(
            buffer, 
            media_type="image/png", 
            headers={"Content-Disposition": f'attachment; filename="{safe_filename}"'}
        )

    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"QR code generation failed: {str(e)}")
    
    